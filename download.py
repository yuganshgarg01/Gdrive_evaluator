"""
download.py
-----------
Downloads all student assignment files from Google Drive
and creates results.xlsx with local file paths.

KEY FEATURE: Saves results.xlsx after every student download,
so you can run evaluate.py even while downloads are still in progress.

Usage:
    python download.py

Requirements:
    pip install gdown pandas openpyxl
"""

import os
import re
import time
import pandas as pd
import gdown
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================

CSV_FILE                = "submissions.csv"
DOWNLOAD_DIR            = "downloads"
OUTPUT_EXCEL            = "results.xlsx"
NUM_ASSIGNMENTS         = 3
DELAY_BETWEEN_DOWNLOADS = 1.5   # seconds

# ============================================================
# HELPERS
# ============================================================

def extract_drive_id(url: str):
    patterns = [
        r"/file/d/([a-zA-Z0-9_-]+)",
        r"id=([a-zA-Z0-9_-]+)",
        r"/open\?id=([a-zA-Z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None


def safe_filename(name: str) -> str:
    return re.sub(r"[^\w\s-]", "", name).strip().replace(" ", "_")


def find_col(df, candidates):
    for c in candidates:
        for col in df.columns:
            if col.strip().lower() == c.lower():
                return col
    return None


def download_file(drive_url: str, dest_path: str) -> bool:
    file_id = extract_drive_id(drive_url)
    if not file_id:
        print(f"    ✗ Could not extract file ID from: {drive_url}")
        return False

    urls_to_try = [
        f"https://drive.google.com/uc?id={file_id}&export=download",
        f"https://drive.google.com/uc?id={file_id}",
        f"https://drive.google.com/file/d/{file_id}/view",
    ]

    for url in urls_to_try:
        try:
            gdown.download(url, dest_path, quiet=True)
            if os.path.exists(dest_path) and os.path.getsize(dest_path) > 100:
                print(f"    ✓ Downloaded → {os.path.basename(dest_path)}")
                return True
            if os.path.exists(dest_path):
                os.remove(dest_path)
        except TypeError:
            try:
                gdown.download(url, dest_path)
                if os.path.exists(dest_path) and os.path.getsize(dest_path) > 100:
                    print(f"    ✓ Downloaded → {os.path.basename(dest_path)}")
                    return True
                if os.path.exists(dest_path):
                    os.remove(dest_path)
            except Exception as e:
                continue
        except Exception as e:
            continue

    print(f"    ✗ All download attempts failed for file ID: {file_id}")
    return False


def guess_extension(path: str) -> str:
    if not os.path.exists(path):
        return ".unknown"
    with open(path, "rb") as f:
        header = f.read(8)
    if header[:4] == b"%PDF":
        return ".pdf"
    if header[:4] == b"PK\x03\x04":
        return ".docx"
    return ".bin"


# ============================================================
# EXCEL HELPERS
# ============================================================

HEADERS = [
    "Timestamp", "Email", "Name", "Course", "Roll Number", "JLU ID",
    "A1 Local Path", "A1 Status", "A1 Marks", "A1 Feedback",
    "A2 Local Path", "A2 Status", "A2 Marks", "A2 Feedback",
    "A3 Local Path", "A3 Status", "A3 Marks", "A3 Feedback",
    "Total Marks"
]

HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR_FILL    = PatternFill("solid", fgColor="2F5496")
CENTER_ALIGN= Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
FILL_OK     = PatternFill("solid", fgColor="E2EFDA")
FILL_FAIL   = PatternFill("solid", fgColor="FCE4D6")
FILL_NA     = PatternFill("solid", fgColor="FFF2CC")
FILL_PENDING= PatternFill("solid", fgColor="DDEBF7")   # blue = pending


def create_excel(output_path: str):
    """Create a fresh results.xlsx with just the header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(HEADERS)

    for col_idx, col_name in enumerate(HEADERS, start=1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    ws.freeze_panes    = "A2"
    ws.row_dimensions[1].height = 30
    wb.save(output_path)
    print(f"  ✓ Created {output_path}")
    return wb


def append_student_row(output_path: str, record: dict):
    """Append one student row to results.xlsx and save immediately."""
    wb = load_workbook(output_path)
    ws = wb.active

    row_data = [record.get(h, "") for h in HEADERS]
    ws.append(row_data)

    # Style the new row
    row_idx = ws.max_row
    for col_idx, key in enumerate(HEADERS, start=1):
        cell           = ws.cell(row=row_idx, column=col_idx)
        cell.border    = THIN_BORDER
        cell.alignment = LEFT_ALIGN
        cell.font      = Font(name="Arial", size=10)

        # Color-code status cells
        if key.endswith("Status"):
            val = str(record.get(key, ""))
            if val == "downloaded":
                cell.fill = FILL_OK
            elif val == "download failed":
                cell.fill = FILL_FAIL
            elif val == "pending":
                cell.fill = FILL_PENDING
            else:
                cell.fill = FILL_NA

    # Auto-widen columns
    for col_idx, key in enumerate(HEADERS, start=1):
        val = str(record.get(key, ""))
        current_width = ws.column_dimensions[get_column_letter(col_idx)].width or 20
        new_width     = min(max(current_width, len(val) + 3), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = new_width

    wb.save(output_path)


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 55)
    print("  📥 Assignment Downloader")
    print("=" * 55)

    if not os.path.exists(CSV_FILE):
        print(f"ERROR: '{CSV_FILE}' not found.")
        return

    df = pd.read_csv(CSV_FILE)
    df.columns = [c.strip() for c in df.columns]
    print(f"\nLoaded {len(df)} student records from {CSV_FILE}")

    col_name      = find_col(df, ["Name", "Student Name", "name"])
    col_roll      = find_col(df, ["Rollnumber", "Roll Number", "Roll no", "RollNo", "rollnumber"])
    col_jlu       = find_col(df, ["JLU id", "JLU ID", "jlu id"])
    col_email     = find_col(df, ["Email Address", "Email", "email"])
    col_course    = find_col(df, ["Course", "course"])
    col_timestamp = find_col(df, ["Timestamp", "timestamp"])

    if not col_name or not col_roll:
        print("ERROR: Could not find Name/Roll columns.")
        print(f"Columns: {df.columns.tolist()}")
        return

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    # Create fresh results.xlsx upfront with just headers
    print(f"\nInitializing {OUTPUT_EXCEL}...")
    if os.path.exists(OUTPUT_EXCEL):
        os.remove(OUTPUT_EXCEL)
    create_excel(OUTPUT_EXCEL)

    print(f"\n{'─'*55}")
    print(f"{'NOTE: results.xlsx is saved after EVERY student.'}")
    print(f"{'You can run evaluate.py at any time in parallel.'}")
    print(f"{'─'*55}\n")

    downloaded_count = 0
    failed_count     = 0
    skipped_count    = 0

    for idx, row in df.iterrows():
        name      = str(row.get(col_name,  f"Student{idx}")).strip()
        rollno    = str(row.get(col_roll,  f"ROLL{idx}")).strip()
        jlu_id    = str(row.get(col_jlu,   "")).strip() if col_jlu    else ""
        email     = str(row.get(col_email, "")).strip() if col_email  else ""
        course    = str(row.get(col_course,"")).strip() if col_course else ""
        timestamp = str(row.get(col_timestamp,"")).strip() if col_timestamp else ""

        print(f"[{idx+1}/{len(df)}] {name} ({rollno})")

        record = {
            "Timestamp":   timestamp,
            "Email":       email,
            "Name":        name,
            "Course":      course,
            "Roll Number": rollno,
            "JLU ID":      jlu_id,
            "Total Marks": ""
        }

        for a_num in range(1, NUM_ASSIGNMENTS + 1):
            # Detect assignment URL column
            drive_url = ""
            for col_try in [
                f"Upload Assignment {a_num}",
                f"Assignment {a_num}",
                f"upload assignment {a_num}",
                f"A{a_num}",
            ]:
                if col_try in df.columns:
                    drive_url = str(row.get(col_try, "")).strip()
                    break

            local_path = ""
            status     = "not submitted"

            if drive_url and drive_url.lower() not in ("nan", "", "none"):
                safe_name  = safe_filename(name)
                temp_path  = os.path.join(DOWNLOAD_DIR, f"{rollno}_{safe_name}_a{a_num}_temp")
                final_base = os.path.join(DOWNLOAD_DIR, f"{rollno}_{safe_name}_a{a_num}")

                success = download_file(drive_url, temp_path)

                if success:
                    ext        = guess_extension(temp_path)
                    final_path = final_base + ext
                    if os.path.exists(final_path):
                        os.remove(final_path)
                    os.rename(temp_path, final_path)
                    local_path = os.path.abspath(final_path)
                    status     = "downloaded"
                    downloaded_count += 1
                else:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    status = "download failed"
                    failed_count += 1

                time.sleep(DELAY_BETWEEN_DOWNLOADS)
            else:
                print(f"    - Assignment {a_num}: not submitted")
                skipped_count += 1

            record[f"A{a_num} Local Path"] = local_path
            record[f"A{a_num} Status"]     = status
            record[f"A{a_num} Marks"]      = ""
            record[f"A{a_num} Feedback"]   = ""

        # ── SAVE IMMEDIATELY after each student ──────────────
        append_student_row(OUTPUT_EXCEL, record)
        print(f"    💾 Saved to {OUTPUT_EXCEL}  "
              f"(downloaded: {downloaded_count}, failed: {failed_count})")

    # ── Final summary ──────────────────────────────────────
    print(f"\n{'='*55}")
    print(f"✓ All done!")
    print(f"  Downloaded     : {downloaded_count}")
    print(f"  Failed         : {failed_count}")
    print(f"  Not submitted  : {skipped_count}")
    print(f"  Results file   : {OUTPUT_EXCEL}")
    print(f"{'='*55}")
    print(f"\nRun evaluate.py now:  python evaluate.py")


if __name__ == "__main__":
    main()